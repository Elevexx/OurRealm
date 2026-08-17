"""Responsibility Center — Bundle F export engine + birthday auto-events.

Async report runs (claim-locked, idempotent), CSV (formula-injection
safe), XLSX (openpyxl), PDF (reportlab). Files stored in Mongo with
time-limited tokened downloads; run audit history outlives files.
"""
import asyncio
import csv
import io
import logging
import uuid
from datetime import datetime, timedelta, timezone

from fastapi import HTTPException
from pymongo.errors import DuplicateKeyError

from core.db import db
from services import responsibility_center as rc
from services import rc_reports
from services.rc_units import _ctx

log = logging.getLogger("ourrealm.rc.exports")

FORMATS = ("csv", "xlsx", "pdf")
FILE_TTL_HOURS = 48
_IDX = False


def _iso(dt=None):
    return (dt or datetime.now(timezone.utc)).isoformat()


async def ensure_export_indexes():
    global _IDX
    if _IDX:
        return
    try:
        await db.responsibility_center_report_runs.create_index(
            [("center_id", 1), ("idempotency_key", 1)], unique=True, name="uniq_run_key",
            partialFilterExpression={"idempotency_key": {"$exists": True}})
        await db.responsibility_center_report_runs.create_index([("status", 1)], name="run_status")
        await db.responsibility_center_birthday_events.create_index(
            [("center_id", 1), ("user_id", 1), ("year", 1)], unique=True, name="uniq_bday")
    except Exception as e:  # noqa: BLE001
        log.warning(f"[rc-exports] index issue: {e}")
    _IDX = True


def _safe_cell(v):
    s = "" if v is None else str(v)
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _build_csv(report: dict, footer: str = "") -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([_safe_cell(f"{report['center_name']} — {report['name']}")])
    w.writerow([_safe_cell(f"Range (UTC): {report['filters']['date_from'][:10]} to {report['filters']['date_to'][:10]} · Generated {report['generated_at'][:16]}Z")])
    w.writerow([])
    w.writerow(["Summary metric", "Value"])
    for k, v in (report.get("summary") or {}).items():
        w.writerow([_safe_cell(k.replace("_", " ")), _safe_cell(v)])
    for bname, rows in (report.get("breakdowns") or {}).items():
        if not rows:
            continue
        w.writerow([])
        w.writerow([_safe_cell(bname.replace("_", " "))])
        for r in rows:
            w.writerow([_safe_cell(r.get("label") or r.get("key")), _safe_cell(r.get("count")),
                        _safe_cell(r.get("fire_power", ""))])
    cols = report.get("columns") or []
    if cols and report.get("rows"):
        w.writerow([])
        w.writerow([_safe_cell(c.replace("_", " ")) for c in cols])
        for r in report["rows"]:
            w.writerow([_safe_cell(r.get(c)) for c in cols])
    if footer:
        w.writerow([])
        w.writerow([_safe_cell(footer)])
    return buf.getvalue().encode("utf-8-sig")


def _build_xlsx(report: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([_safe_cell(f"{report['center_name']} — {report['name']}")])
    ws.append([_safe_cell(f"Range (UTC): {report['filters']['date_from'][:10]} to {report['filters']['date_to'][:10]}")])
    ws.append([_safe_cell(f"Generated {report['generated_at'][:16]}Z")])
    ws.append([])
    ws.append(["Metric", "Value"])
    for k, v in (report.get("summary") or {}).items():
        ws.append([_safe_cell(k.replace("_", " ")), _safe_cell(v)])
    for bname, rows in (report.get("breakdowns") or {}).items():
        if not rows:
            continue
        ws.append([])
        ws.append([_safe_cell(bname.replace("_", " ")), "count"])
        for r in rows:
            ws.append([_safe_cell(r.get("label") or r.get("key")), r.get("count")])
    cols = report.get("columns") or []
    if cols and report.get("rows"):
        ds = wb.create_sheet("Data")
        ds.append([_safe_cell(c.replace("_", " ")) for c in cols])
        ds.freeze_panes = "A2"
        for r in report["rows"]:
            ds.append([_safe_cell(r.get(c)) for c in cols])
        for i, c in enumerate(cols, 1):
            ds.column_dimensions[get_column_letter(i)].width = max(14, min(len(c) + 6, 34))
    if footer:
        ws.append([])
        ws.append([_safe_cell(footer)])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_pdf(report: dict, requested_by: str, footer: str = "") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    def _esc(v):
        return str("" if v is None else v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")[:180]
    buf = io.BytesIO()

    def footer_fn(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 6)
        note = footer or "For authorized Center members."
        for i, chunk in enumerate([note[j:j + 150] for j in range(0, len(note), 150)][:3]):
            canvas.drawString(0.6 * inch, 0.62 * inch - i * 0.11 * inch, chunk)
        canvas.drawString(0.6 * inch, 0.28 * inch,
                          f"OurRealm Responsibility Center · Page {doc.page}")
        canvas.restoreState()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.7 * inch)
    st = getSampleStyleSheet()
    story = [Paragraph(_esc(report["center_name"]), st["Title"]),
             Paragraph(_esc(report["name"]), st["Heading2"]),
             Paragraph(f"Range (UTC): {_esc(report['filters']['date_from'][:10])} to {_esc(report['filters']['date_to'][:10])}"
                       f" · Generated {_esc(report['generated_at'][:16])}Z · By @{_esc(requested_by)}", st["Normal"]),
             Spacer(1, 10)]
    summary = [["Metric", "Value"]] + [[_esc(k.replace("_", " ")), _esc(v)]
                                       for k, v in (report.get("summary") or {}).items()]
    t = Table(summary, hAlign="LEFT", colWidths=[3.4 * inch, 3.4 * inch])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1b2838")),
                           ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                           ("FONTSIZE", (0, 0), (-1, -1), 8),
                           ("GRID", (0, 0), (-1, -1), 0.4, colors.grey)]))
    story += [t, Spacer(1, 10)]
    chart = _pdf_chart(report)
    if chart is not None:
        story += [chart, Spacer(1, 6)]
    for bname, rows in (report.get("breakdowns") or {}).items():
        if not rows:
            continue
        story.append(Paragraph(_esc(bname.replace("_", " ").title()), st["Heading4"]))
        bt = Table([["Group", "Count"]] + [[_esc(r.get("label") or r.get("key")), _esc(r.get("count"))] for r in rows[:25]],
                   hAlign="LEFT", colWidths=[4.4 * inch, 2.4 * inch])
        bt.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 8),
                                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eef5"))]))
        story += [bt, Spacer(1, 8)]
    cols = report.get("columns") or []
    if cols and report.get("rows"):
        story.append(Paragraph("Detail", st["Heading4"]))
        data = [[_esc(c.replace("_", " ")) for c in cols]] + \
               [[_esc(r.get(c)) for c in cols] for r in report["rows"][:400]]
        dt = Table(data, hAlign="LEFT", repeatRows=1)
        dt.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 6.5),
                                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1b2838")),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white)]))
        story.append(dt)
    if report.get("note"):
        story += [Spacer(1, 8), Paragraph(_esc(report["note"]), st["Italic"])]
    doc.build(story, onFirstPage=footer_fn, onLaterPages=footer_fn)
    return buf.getvalue()


MIME = {"csv": "text/csv; charset=utf-8",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf"}


def _pdf_chart(report: dict):
    """Rendered bar chart from real breakdown data; None on any failure."""
    try:
        from reportlab.graphics.shapes import Drawing, String
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib import colors
        bname, rows = next(((k, v) for k, v in (report.get("breakdowns") or {}).items() if v), (None, None))
        if not rows:
            return None
        rows = rows[:8]
        d = Drawing(440, 190)
        chart = VerticalBarChart()
        chart.x, chart.y, chart.width, chart.height = 30, 30, 390, 130
        chart.data = [[r.get("count") or 0 for r in rows]]
        chart.categoryAxis.categoryNames = [str(r.get("label") or r.get("key"))[:14] for r in rows]
        chart.categoryAxis.labels.fontSize = 6
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor("#2f6db3")
        d.add(chart)
        d.add(String(30, 172, f"{bname.replace('_', ' ').title()} (chart of the table below)", fontSize=8))
        return d
    except Exception:  # noqa: BLE001
        return None


# ── Report runs ──────────────────────────────────────────────────────────
async def create_run(user: dict, center_id: str, body: dict) -> dict:
    await ensure_export_indexes()
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    if "export_reports" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to export reports")
    report_key = body.get("report_key")
    meta = rc_reports.REPORTS.get(report_key)
    if not meta:
        raise HTTPException(status_code=404, detail="Unknown report")
    if meta["perm"] not in perms:
        raise HTTPException(status_code=403, detail="You don't have access to this report")
    fmt = body.get("format") or "csv"
    if fmt not in FORMATS:
        raise HTTPException(status_code=400, detail="Invalid format")
    recent = await db.responsibility_center_report_runs.count_documents(
        {"requested_by": user["id"], "requested_at": {"$gte": (datetime.now(timezone.utc) - timedelta(minutes=10)).isoformat()}})
    if recent >= 15:
        raise HTTPException(status_code=429, detail="Too many exports — please wait a few minutes")
    run = {"id": uuid.uuid4().hex, "center_id": center_id, "report_key": report_key,
           "report_name": meta["name"], "format": fmt,
           "filters": rc_reports.parse_filters(body.get("filters") or {}),
           "requested_by": user["id"], "requested_by_username": user.get("username"),
           "status": "queued", "progress_percent": 0, "record_count": None,
           "failure_reason": None, "download_token": uuid.uuid4().hex,
           "requested_at": _iso(), "started_at": None, "completed_at": None,
           "expires_at": None, "downloaded_at": None,
           "created_at": _iso(), "updated_at": _iso()}
    if body.get("client_token"):
        run["idempotency_key"] = f"{report_key}:{fmt}:{str(body['client_token'])[:60]}"
    try:
        await db.responsibility_center_report_runs.insert_one({**run})
    except DuplicateKeyError:
        existing = await db.responsibility_center_report_runs.find_one(
            {"center_id": center_id, "idempotency_key": run["idempotency_key"]},
            {"_id": 0, "download_token": 0})
        return {"run": existing, "duplicate": True}
    asyncio.create_task(run_export_pass())
    await rc.log_activity(center_id, user, "report_export_requested",
                          f"@{user.get('username')} requested a {fmt.upper()} export of \"{meta['name']}\"")
    return {"run": {k: v for k, v in run.items() if k != "download_token"}}


async def run_export_pass() -> dict:
    await ensure_export_indexes()
    now = datetime.now(timezone.utc)
    done = 0
    while done < 20:
        run = await db.responsibility_center_report_runs.find_one_and_update(
            {"status": "queued",
             "$or": [{"claim_until": None}, {"claim_until": {"$exists": False}},
                     {"claim_until": {"$lt": now.isoformat()}}]},
            {"$set": {"status": "processing", "started_at": _iso(),
                      "claim_until": (now + timedelta(minutes=5)).isoformat(),
                      "updated_at": _iso()}},
            projection={"_id": 0})
        if not run:
            break
        done += 1
        try:
            user = await db.users.find_one({"id": run["requested_by"]}, {"_id": 0, "id": 1, "username": 1})
            report = await rc_reports.run_report(user, run["center_id"], run["report_key"], run["filters"])
            comp = await get_compliance_settings()
            footer = comp["export_footer_text"] if comp["export_footer_enabled"] else ""
            if run["format"] == "csv":
                data = _build_csv(report, footer)
            elif run["format"] == "xlsx":
                data = _build_xlsx(report, footer)
            else:
                data = _build_pdf(report, run.get("requested_by_username") or "", footer)
            expires = (now + timedelta(hours=FILE_TTL_HOURS)).isoformat()
            await db.responsibility_center_report_files.update_one(
                {"run_id": run["id"]},
                {"$set": {"run_id": run["id"], "center_id": run["center_id"], "data": data,
                          "mime": MIME[run["format"]], "size": len(data),
                          "filename": f"{run['report_key']}-{run['requested_at'][:10]}.{run['format']}",
                          "expires_at": expires, "created_at": _iso()}}, upsert=True)
            await db.responsibility_center_report_runs.update_one(
                {"id": run["id"]},
                {"$set": {"status": "ready", "progress_percent": 100,
                          "record_count": len(report.get("rows") or []),
                          "completed_at": _iso(), "expires_at": expires,
                          "claim_until": None, "updated_at": _iso()}})
            await rc.notify_user(run["requested_by"], "responsibility_center_report_ready",
                                 f"Your \"{run['report_name']}\" {run['format'].upper()} export is ready to download.",
                                 f"/responsibility-center/{run['center_id']}?tab=reports&run={run['id']}",
                                 run["center_id"])
        except Exception as e:  # noqa: BLE001
            log.exception("[rc-exports] run %s failed", run["id"])
            await db.responsibility_center_report_runs.update_one(
                {"id": run["id"]},
                {"$set": {"status": "failed", "failure_reason": str(e)[:300],
                          "claim_until": None, "updated_at": _iso()}})
            await rc.notify_user(run["requested_by"], "responsibility_center_report_failed",
                                 f"Your \"{run.get('report_name')}\" export failed — you can retry from Export History.",
                                 f"/responsibility-center/{run['center_id']}?tab=reports",
                                 run["center_id"])
    expired = await db.responsibility_center_report_runs.update_many(
        {"status": "ready", "expires_at": {"$lt": now.isoformat()}},
        {"$set": {"status": "expired", "updated_at": _iso()}})
    if expired.modified_count:
        await db.responsibility_center_report_files.delete_many(
            {"expires_at": {"$lt": now.isoformat()}})
    return {"exports_processed": done, "expired": expired.modified_count}


async def list_runs(user: dict, center_id: str) -> dict:
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    q = {"center_id": center_id}
    if "manage_roles" not in perms:  # non-admins see their own exports
        q["requested_by"] = user["id"]
    runs = await db.responsibility_center_report_runs.find(
        q, {"_id": 0, "download_token": 0}).sort("requested_at", -1).to_list(50)
    return {"runs": runs}


async def get_download(user: dict, center_id: str, run_id: str):
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    run = await db.responsibility_center_report_runs.find_one(
        {"id": run_id, "center_id": center_id}, {"_id": 0})
    if not run:
        raise HTTPException(status_code=404, detail="Export not found")
    if run["requested_by"] != user["id"] and "manage_roles" not in perms:
        raise HTTPException(status_code=403, detail="You can't download this export")
    meta = rc_reports.REPORTS.get(run["report_key"]) or {}
    if meta.get("perm") and meta["perm"] not in perms:
        raise HTTPException(status_code=403, detail="You no longer have access to this report")
    if run["status"] != "ready" or (run.get("expires_at") and run["expires_at"] < _iso()):
        raise HTTPException(status_code=410, detail="This export has expired — generate it again")
    f = await db.responsibility_center_report_files.find_one({"run_id": run_id})
    if not f:
        raise HTTPException(status_code=410, detail="This export file is no longer available")
    await db.responsibility_center_report_runs.update_one(
        {"id": run_id, "downloaded_at": None}, {"$set": {"downloaded_at": _iso()}})
    return f["data"], f["mime"], f["filename"]


async def retry_run(user: dict, center_id: str, run_id: str) -> dict:
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    if "export_reports" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to export reports")
    upd = await db.responsibility_center_report_runs.update_one(
        {"id": run_id, "center_id": center_id, "requested_by": user["id"],
         "status": {"$in": ["failed", "expired"]}},
        {"$set": {"status": "queued", "failure_reason": None, "claim_until": None,
                  "requested_at": _iso(), "updated_at": _iso()}})
    if upd.modified_count != 1:
        raise HTTPException(status_code=409, detail="Only your failed or expired exports can be retried")
    asyncio.create_task(run_export_pass())
    return {"ok": True}


# ── Birthday auto-events (opt-in, privacy controlled) ────────────────────
async def get_birthday_settings(user: dict, center_id: str) -> dict:
    center, membership, perms = await _ctx(center_id, user, "view_units", write=False)
    consent = await db.responsibility_center_birthday_consents.find_one(
        {"center_id": center_id, "user_id": user["id"]}, {"_id": 0})
    return {"birthday_auto_events_enabled": bool(center.get("birthday_auto_events_enabled")),
            "birthday_show_year": bool(center.get("birthday_show_year")),
            "can_manage": "edit_center" in perms,
            "my_consent": bool(consent and consent.get("consented")),
            "my_birth_month": (consent or {}).get("birth_month"),
            "my_birth_day": (consent or {}).get("birth_day")}


async def update_birthday_settings(user: dict, center_id: str, body: dict) -> dict:
    center, membership, perms = await _ctx(center_id, user, "edit_center")
    sets = {}
    for k in ("birthday_auto_events_enabled", "birthday_show_year"):
        if k in body:
            sets[k] = bool(body[k])
    if sets:
        await db.responsibility_centers.update_one({"id": center_id}, {"$set": sets})
        await rc.log_activity(center_id, user, "birthday_settings_changed",
                              f"@{user.get('username')} updated birthday auto-event settings")
    return await get_birthday_settings(user, center_id)


async def set_birthday_consent(user: dict, center_id: str, body: dict) -> dict:
    await ensure_export_indexes()
    await _ctx(center_id, user, "", write=False)
    consented = bool(body.get("consented"))
    doc = {"center_id": center_id, "user_id": user["id"], "consented": consented,
           "updated_at": _iso()}
    if consented:
        month, day = int(body.get("birth_month") or 0), int(body.get("birth_day") or 0)
        if not (1 <= month <= 12 and 1 <= day <= 31):
            raise HTTPException(status_code=400, detail="Provide a valid birth month and day to opt in")
        doc.update(birth_month=month, birth_day=day,
                   birth_year=int(body["birth_year"]) if body.get("birth_year") else None)
    await db.responsibility_center_birthday_consents.update_one(
        {"center_id": center_id, "user_id": user["id"]}, {"$set": doc}, upsert=True)
    if not consented:
        await db.responsibility_center_calendar_events.update_many(
            {"center_id": center_id, "birthday_user_id": user["id"], "status": "scheduled"},
            {"$set": {"status": "canceled", "canceled_at": _iso(), "updated_at": _iso()}})
        await db.responsibility_center_birthday_events.delete_many(
            {"center_id": center_id, "user_id": user["id"]})
    return {"ok": True, "consented": consented}


async def run_birthday_pass() -> dict:
    """Yearly birthday events for consented members in opted-in Centers.
    Unique (center, user, year) claim — concurrency-safe."""
    await ensure_export_indexes()
    now = datetime.now(timezone.utc)
    created = 0
    async for center in db.responsibility_centers.find(
            {"birthday_auto_events_enabled": True, "status": "active"},
            {"_id": 0, "id": 1, "name": 1, "birthday_show_year": 1}):
        async for consent in db.responsibility_center_birthday_consents.find(
                {"center_id": center["id"], "consented": True}, {"_id": 0}):
            m = await db.responsibility_center_memberships.find_one(
                {"center_id": center["id"], "user_id": consent["user_id"], "status": "active"},
                {"_id": 0, "user_id": 1})
            if not m:
                continue
            try:
                next_bday = datetime(now.year, consent["birth_month"], min(consent["birth_day"], 28)
                                     if (consent["birth_month"], consent["birth_day"]) == (2, 29) else consent["birth_day"],
                                     tzinfo=timezone.utc)
            except ValueError:
                continue
            if next_bday < now - timedelta(days=1):
                next_bday = next_bday.replace(year=now.year + 1)
            if next_bday > now + timedelta(days=180):
                continue
            year = next_bday.year
            try:
                await db.responsibility_center_birthday_events.insert_one(
                    {"center_id": center["id"], "user_id": consent["user_id"], "year": year,
                     "created_at": _iso()})
            except DuplicateKeyError:
                continue
            u = await db.users.find_one({"id": consent["user_id"]}, {"_id": 0, "username": 1, "name": 1})
            title = f"{(u or {}).get('name') or '@' + ((u or {}).get('username') or 'Member')}'s Birthday"
            await db.responsibility_center_calendar_events.insert_one({
                "id": uuid.uuid4().hex, "center_id": center["id"], "unit_id": None,
                "event_type": "birthday", "title": title, "description": "",
                "visibility": "center", "created_by": consent["user_id"],
                "created_by_username": (u or {}).get("username"),
                "organizer_id": consent["user_id"], "birthday_user_id": consent["user_id"],
                "start_at": next_bday.isoformat(),
                "end_at": (next_bday + timedelta(hours=23, minutes=59)).isoformat(),
                "all_day": True, "timezone": "UTC", "location": "", "virtual_link": None,
                "status": "scheduled", "attendance_enabled": False,
                "reminders": [1440], "attendees": [], "related_item_id": None,
                "version": 1, "created_at": _iso(), "updated_at": _iso(), "canceled_at": None})
            created += 1
    return {"birthday_events_created": created}


# ── Scheduled report delivery (Bundle G — explicit opt-in only) ──────────
SCHED_FREQS = ("weekly", "monthly")


def _next_run(freq: str, now=None):
    now = now or datetime.now(timezone.utc)
    return (now + (timedelta(days=7) if freq == "weekly" else timedelta(days=30))).isoformat()


def _next_run_for(sched: dict, now=None) -> str:
    """Next occurrence honouring day / hour / timezone, returned as UTC iso."""
    from zoneinfo import ZoneInfo
    now = now or datetime.now(timezone.utc)
    try:
        tz = ZoneInfo(sched.get("timezone") or "UTC")
    except Exception:  # noqa: BLE001
        tz = timezone.utc
    hour = sched.get("send_hour")
    hour = 8 if hour is None else min(max(int(hour), 0), 23)
    local = now.astimezone(tz)
    if (sched.get("frequency") or "weekly") == "weekly":
        dow = sched.get("day_of_week")
        dow = 0 if dow is None else min(max(int(dow), 0), 6)
        cand = (local + timedelta(days=(dow - local.weekday()) % 7)).replace(
            hour=hour, minute=0, second=0, microsecond=0)
        if cand <= local:
            cand += timedelta(days=7)
    else:
        dom = sched.get("day_of_month")
        dom = 1 if dom is None else min(max(int(dom), 1), 28)
        cand = local.replace(day=dom, hour=hour, minute=0, second=0, microsecond=0)
        if cand <= local:
            cand = (cand.replace(day=1) + timedelta(days=32)).replace(day=dom)
    return cand.astimezone(timezone.utc).isoformat()


async def create_schedule(user: dict, center_id: str, body: dict) -> dict:
    from services import rc_reports
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    if "export_reports" not in perms:
        raise HTTPException(status_code=403, detail="You don't have permission to schedule reports")
    meta = rc_reports.REPORTS.get(body.get("report_key"))
    if not meta:
        raise HTTPException(status_code=404, detail="Unknown report")
    if meta["perm"] not in perms:
        raise HTTPException(status_code=403, detail="You don't have access to this report")
    freq = body.get("frequency") or "weekly"
    if freq not in SCHED_FREQS:
        raise HTTPException(status_code=400, detail="Frequency must be weekly or monthly")
    fmt = body.get("format") or "csv"
    if fmt not in FORMATS:
        raise HTTPException(status_code=400, detail="Invalid format")
    from zoneinfo import ZoneInfo
    tzname = (body.get("timezone") or "UTC").strip()
    try:
        ZoneInfo(tzname)
    except Exception:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="Invalid timezone")
    sched = {"id": uuid.uuid4().hex, "center_id": center_id, "report_key": body["report_key"],
             "report_name": meta["name"], "frequency": freq, "format": fmt,
             "filters": rc_reports.parse_filters(body.get("filters") or {}),
             "recipient_ids": list(dict.fromkeys(body.get("recipient_ids") or [user["id"]]))[:20],
             "delivery_channel": "in_app",  # email deferred — never silent external delivery
             "enabled": bool(body.get("enabled", False)),  # OFF unless explicitly enabled
             "day_of_week": None if body.get("day_of_week") is None else min(max(int(body["day_of_week"]), 0), 6),
             "day_of_month": None if body.get("day_of_month") is None else min(max(int(body["day_of_month"]), 1), 28),
             "send_hour": None if body.get("send_hour") is None else min(max(int(body["send_hour"]), 0), 23),
             "timezone": tzname,
             "owner_id": user["id"], "owner_username": user.get("username"),
             "last_run_at": None, "failure_count": 0,
             "created_at": _iso(), "updated_at": _iso()}
    sched["next_run_at"] = _next_run_for(sched) if sched["enabled"] else None
    await db.responsibility_center_scheduled_reports.insert_one({**sched})
    await rc.log_activity(center_id, user, "scheduled_report_created",
                          f"@{user.get('username')} scheduled the {meta['name']} report ({freq})")
    return {"schedule": sched}


async def list_schedules(user: dict, center_id: str) -> dict:
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    rows = await db.responsibility_center_scheduled_reports.find(
        {"center_id": center_id}, {"_id": 0}).sort("created_at", -1).to_list(30)
    return {"schedules": rows}


async def update_schedule(user: dict, center_id: str, schedule_id: str, body: dict) -> dict:
    center, membership, perms = await _ctx(center_id, user, "view_reports", write=False)
    if "export_reports" not in perms:
        raise HTTPException(status_code=403, detail="You can't manage report schedules")
    sched = await db.responsibility_center_scheduled_reports.find_one(
        {"id": schedule_id, "center_id": center_id}, {"_id": 0})
    if not sched:
        raise HTTPException(status_code=404, detail="Schedule not found")
    if body.get("delete"):
        await db.responsibility_center_scheduled_reports.delete_one(
            {"id": schedule_id, "center_id": center_id})
        # audit history (activity log + run claims) is preserved after deletion
        await rc.log_activity(center_id, user, "scheduled_report_deleted",
                              f"@{user.get('username')} deleted the {sched['report_name']} schedule")
        return {"ok": True, "deleted": True}
    from zoneinfo import ZoneInfo
    sets = {"updated_at": _iso()}
    if "frequency" in body:
        if body["frequency"] not in SCHED_FREQS:
            raise HTTPException(status_code=400, detail="Frequency must be weekly or monthly")
        sets["frequency"] = body["frequency"]
    if "format" in body:
        if body["format"] not in FORMATS:
            raise HTTPException(status_code=400, detail="Invalid format")
        sets["format"] = body["format"]
    if "timezone" in body:
        tzname = (body["timezone"] or "UTC").strip()
        try:
            ZoneInfo(tzname)
        except Exception:  # noqa: BLE001
            raise HTTPException(status_code=400, detail="Invalid timezone")
        sets["timezone"] = tzname
    for f, lo, hi in (("day_of_week", 0, 6), ("day_of_month", 1, 28), ("send_hour", 0, 23)):
        if f in body:
            sets[f] = None if body[f] is None else min(max(int(body[f]), lo), hi)
    if "recipient_ids" in body:
        sets["recipient_ids"] = list(dict.fromkeys(body["recipient_ids"] or [user["id"]]))[:20]
    if "filters" in body:
        from services import rc_reports
        sets["filters"] = rc_reports.parse_filters(body["filters"] or {})
    if "enabled" in body:
        sets["enabled"] = bool(body["enabled"])
        sets["failure_count"] = 0
    merged = {**sched, **sets}
    if merged.get("enabled"):
        sets["next_run_at"] = _next_run_for(merged)
    elif "enabled" in body:
        sets["next_run_at"] = None
    await db.responsibility_center_scheduled_reports.update_one(
        {"id": schedule_id, "center_id": center_id}, {"$set": sets})
    return {"ok": True}


async def run_scheduled_reports_pass() -> dict:
    """Claim-locked, idempotent per period — one run per schedule period
    even under overlapping workers. Permissions rechecked at run time."""
    from services import rc_reports
    now = datetime.now(timezone.utc)
    ran = 0
    try:
        await db.responsibility_center_scheduled_report_runs.create_index(
            [("period_key", 1)], unique=True, name="uniq_sched_period")
    except Exception:  # noqa: BLE001
        pass
    async for sched in db.responsibility_center_scheduled_reports.find(
            {"enabled": True, "next_run_at": {"$lte": now.isoformat()}}, {"_id": 0}).limit(25):
        period_key = f"{sched['id']}:{(sched.get('next_run_at') or now.isoformat())[:10]}"
        try:
            await db.responsibility_center_scheduled_report_runs.insert_one(
                {"period_key": period_key, "schedule_id": sched["id"], "ran_at": _iso()})
        except DuplicateKeyError:
            continue  # another worker claimed this period
        try:
            center = await db.responsibility_centers.find_one({"id": sched["center_id"]}, {"_id": 0})
            owner = await db.users.find_one({"id": sched["owner_id"]}, {"_id": 0, "id": 1, "username": 1})
            if not center or (center.get("status") or "active") != "active" or not owner:
                raise ValueError("Center inactive or owner missing")
            m = await db.responsibility_center_memberships.find_one(
                {"center_id": sched["center_id"], "user_id": sched["owner_id"], "status": "active"}, {"_id": 0, "role": 1})
            perms = set(rc.ROLE_PERMISSIONS.get((m or {}).get("role") or "member", set()))
            meta = rc_reports.REPORTS.get(sched["report_key"]) or {}
            if "export_reports" not in perms or meta.get("perm") not in perms:
                raise ValueError("Owner no longer has report permission")
            filters = dict(sched.get("filters") or {})
            span = 7 if sched["frequency"] == "weekly" else 30
            filters["date_from"] = (now - timedelta(days=span)).isoformat()
            filters["date_to"] = now.isoformat()
            run = await create_run(owner, sched["center_id"],
                                   {"report_key": sched["report_key"], "format": sched["format"],
                                    "filters": filters, "client_token": f"sched:{period_key}"})
            for rid in sched.get("recipient_ids") or []:
                rm = await db.responsibility_center_memberships.find_one(
                    {"center_id": sched["center_id"], "user_id": rid, "status": "active"}, {"_id": 1})
                if rm:  # removed recipients excluded
                    await rc.notify_user(rid, "responsibility_center_scheduled_report",
                                         f"Your scheduled \"{sched['report_name']}\" report is being generated.",
                                         f"/responsibility-center/{sched['center_id']}?tab=reports",
                                         sched["center_id"])
            await db.responsibility_center_scheduled_reports.update_one(
                {"id": sched["id"]},
                {"$set": {"last_run_at": _iso(), "next_run_at": _next_run_for(sched, now),
                          "failure_count": 0, "updated_at": _iso()}})
            ran += 1
        except Exception as e:  # noqa: BLE001
            fails = int(sched.get("failure_count") or 0) + 1
            sets = {"failure_count": fails, "last_failure": str(e)[:200],
                    "next_run_at": _next_run_for(sched, now), "updated_at": _iso()}
            if fails >= 3:
                sets["enabled"] = False  # disabled after repeated failures, with notice
                await rc.notify_user(sched["owner_id"], "responsibility_center_scheduled_report_failed",
                                     f"Your scheduled \"{sched['report_name']}\" report was paused after repeated failures.",
                                     f"/responsibility-center/{sched['center_id']}?tab=reports",
                                     sched["center_id"])
            await db.responsibility_center_scheduled_reports.update_one(
                {"id": sched["id"]}, {"$set": sets})
    return {"scheduled_reports_ran": ran}


# ── Legal & Compliance settings (Bundle G review — admin configurable) ───
DEFAULT_EXPORT_FOOTER = ("This report is generated from information recorded within this "
                         "Responsibility Center. It is provided for organizational and informational "
                         "purposes and may not constitute an official, certified, legally recognized, "
                         "educational, employment, governmental, financial, or medical record.")
COMPLIANCE_DEFAULTS = {
    "export_footer_enabled": True,
    "export_footer_text": DEFAULT_EXPORT_FOOTER,
    "fire_power_note": "Fire Power is an internal engagement resource. It is never money, has no monetary value, and cannot be redeemed or exchanged.",
    "attendance_note": "Attendance tracking within this Center is informational and is not a certified or legally recognized record.",
    "platform_descriptor": "Organizational collaboration and responsibility management platform.",
    "show_disclaimers": True,
}


async def get_compliance_settings() -> dict:
    row = await db.rc_compliance_settings.find_one({"id": "main"}, {"_id": 0}) or {}
    return {**COMPLIANCE_DEFAULTS, **{k: v for k, v in row.items() if k in COMPLIANCE_DEFAULTS}}


async def update_compliance_settings(body: dict) -> dict:
    sets = {}
    for k in COMPLIANCE_DEFAULTS:
        if k in body:
            v = body[k]
            sets[k] = bool(v) if isinstance(COMPLIANCE_DEFAULTS[k], bool) else str(v)[:600]
    if sets:
        await db.rc_compliance_settings.update_one(
            {"id": "main"}, {"$set": sets, "$setOnInsert": {"id": "main"}}, upsert=True)
    return await get_compliance_settings()
